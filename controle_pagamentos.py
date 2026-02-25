"""
Sistema de Controle de Processos de Pagamento — v5.0
Prefeitura Municipal de Governador Edison Lobão/MA — Controladoria Geral

Novidades v5.0:
  • Tema claro e moderno (Light Mode, Inter font)
  • Dashboard com KPIs, gráficos e atividade recente
  • Busca inteligente: exibe recentes sem digitar + filtros por data/status
  • Autopreenchimento CNPJ/CPF e Secretário por histórico
  • Máscaras de entrada (CNPJ, Moeda)
  • Logs de geração de PDF com rastreamento de erros
  • Sincronização nuvem via Google Sheets (opcional)
  • Feedback visual imediato (spinners, confirmações, alertas)
  • UX aprimorada: menos cliques, fluxo linear
"""

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import os, io, json, re, socket, hashlib, secrets, logging
from datetime import date, datetime, timedelta
from copy import deepcopy
import openpyxl
from gerador_pdf import (
    gerar_pdf_capa, gerar_pdf_parecer_padrao,
    gerar_pdf_parecer_passagem, proximo_numero,
)

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING  (rastreamento de erros no PDF e ações do sistema)
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("sistema.log", encoding="utf-8", mode="a"),
    ],
)
log = logging.getLogger("v5")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
_BASE = os.path.dirname(os.path.abspath(__file__))

def cfg_ler() -> dict:
    c = {}
    p = os.path.join(_BASE, "config.txt")
    if os.path.exists(p):
        for ln in open(p, encoding="utf-8"):
            ln = ln.strip()
            if ln and not ln.startswith("#"):
                k, _, v = ln.partition("=")
                c[k.strip()] = v.strip()
    return c

_cfg         = cfg_ler()
EXCEL_PATH   = _cfg.get("PLANILHA",    os.path.join(_BASE, "Controle_Geral_2.xlsx"))
USAR_GSHEETS = _cfg.get("USAR_GSHEETS", "nao").lower() == "sim"

# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Controladoria — Gov. Edison Lobão",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# TEMA CLARO MODERNO
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* Reset base */
html, body, [class*="css"] {
  font-family: 'Inter', 'Segoe UI', sans-serif !important;
  background: #f8fafc !important;
}
.main .block-container { padding: 0 !important; max-width: 100% !important; }

/* Sidebar */
section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #1e3a5f 0%, #1d4ed8 100%) !important;
  min-width: 260px !important;
  box-shadow: 2px 0 16px #1d4ed820 !important;
}
section[data-testid="stSidebar"] * { color: #e0eaff !important; }
section[data-testid="stSidebar"] .stRadio label { font-size: 13px !important; padding: 8px 12px !important; border-radius: 8px !important; transition: background .15s !important; }
section[data-testid="stSidebar"] .stRadio label:hover { background: #ffffff18 !important; }

/* Inputs */
.stTextInput > div > div > input,
.stTextArea > div > textarea,
.stSelectbox > div > div,
.stNumberInput > div > div > input {
  background: #ffffff !important;
  border: 1.5px solid #e2e8f0 !important;
  border-radius: 8px !important;
  color: #1e293b !important;
  font-size: 14px !important;
  transition: border .15s, box-shadow .15s !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > textarea:focus {
  border-color: #2563eb !important;
  box-shadow: 0 0 0 3px #2563eb18 !important;
  outline: none !important;
}
.stTextInput label, .stTextArea label, .stSelectbox label, .stDateInput label, .stNumberInput label {
  font-size: 12px !important; font-weight: 600 !important;
  color: #374151 !important; margin-bottom: 2px !important;
}

/* Buttons */
.stButton > button {
  border-radius: 8px !important; font-weight: 600 !important;
  font-size: 13px !important; transition: all .15s !important;
  border: none !important;
}
.stButton > button[kind="primary"] {
  background: linear-gradient(135deg, #1d4ed8, #3b82f6) !important;
  color: #fff !important; box-shadow: 0 2px 8px #1d4ed830 !important;
}
.stButton > button[kind="primary"]:hover {
  transform: translateY(-1px) !important;
  box-shadow: 0 4px 16px #1d4ed845 !important;
}
.stButton > button[kind="secondary"] {
  background: #ffffff !important; color: #374151 !important;
  border: 1.5px solid #e2e8f0 !important;
}

/* Cards */
.card {
  background: #ffffff;
  border: 1.5px solid #e2e8f0;
  border-radius: 14px;
  padding: 20px 24px;
  margin-bottom: 16px;
  box-shadow: 0 1px 8px #0000000a;
  transition: box-shadow .2s;
}
.card:hover { box-shadow: 0 4px 20px #0000001a; }

/* KPI cards */
.kpi { text-align: center; padding: 22px 16px; border-radius: 14px; border: 1.5px solid #e2e8f0; background: #fff; }
.kpi-val { font-size: 32px; font-weight: 800; color: #0f172a; margin: 6px 0; line-height: 1; }
.kpi-lbl { font-size: 11px; font-weight: 700; color: #64748b; text-transform: uppercase; letter-spacing: .07em; }
.kpi-delta { font-size: 12px; font-weight: 600; margin-top: 4px; }
.kpi-delta.up   { color: #16a34a; }
.kpi-delta.down { color: #dc2626; }

/* Section headers */
.sec-hdr {
  font-size: 10px; font-weight: 700; color: #94a3b8;
  text-transform: uppercase; letter-spacing: .09em;
  margin: 18px 0 8px; display: flex; align-items: center; gap: 6px;
}
.sec-hdr::after { content: ''; flex: 1; height: 1px; background: #f1f5f9; margin-left: 8px; }

/* Status badges */
.badge { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 11px; font-weight: 700; }
.badge-green  { background: #dcfce7; color: #15803d; }
.badge-red    { background: #fee2e2; color: #dc2626; }
.badge-blue   { background: #dbeafe; color: #1d4ed8; }
.badge-orange { background: #ffedd5; color: #c2410c; }
.badge-purple { background: #ede9fe; color: #7c3aed; }
.badge-teal   { background: #ccfbf1; color: #0f766e; }

/* Page header */
.page-hdr {
  background: linear-gradient(135deg, #1e3a5f 0%, #1d4ed8 100%);
  padding: 22px 32px; margin-bottom: 0;
  display: flex; align-items: center; gap: 16px;
}
.page-hdr h2 { color: #ffffff !important; margin: 0; font-size: 20px; font-weight: 800; }
.page-hdr p  { color: #93c5fd; margin: 2px 0 0; font-size: 13px; }

/* Activity feed */
.activity-item {
  display: flex; align-items: flex-start; gap: 12px;
  padding: 12px 0; border-bottom: 1px solid #f1f5f9;
}
.activity-dot { width: 10px; height: 10px; border-radius: 50%; margin-top: 4px; flex-shrink: 0; }

/* Search highlight */
.search-result {
  background: #fff; border: 1.5px solid #e2e8f0; border-radius: 10px;
  padding: 12px 16px; margin: 4px 0; cursor: pointer;
  transition: all .15s;
}
.search-result:hover { border-color: #2563eb; background: #eff6ff; }

/* Log entries */
.log-ok   { color: #16a34a; font-family: monospace; font-size: 12px; }
.log-err  { color: #dc2626; font-family: monospace; font-size: 12px; }
.log-info { color: #2563eb; font-family: monospace; font-size: 12px; }
.log-warn { color: #d97706; font-family: monospace; font-size: 12px; }

/* Field hint */
.field-hint { font-size: 11px; color: #94a3b8; margin-top: 2px; }
.field-auto { font-size: 11px; color: #16a34a; font-weight: 600; margin-top: 2px; }

/* Divider */
.divider { height: 1px; background: #f1f5f9; margin: 16px 0; }

/* Dataframe */
.stDataFrame { border-radius: 10px !important; border: 1.5px solid #e2e8f0 !important; }

/* Progress bar */
.stProgress > div > div { background: #1d4ed8 !important; border-radius: 4px !important; }

/* Checkbox */
.stCheckbox label { font-size: 13px !important; color: #374151 !important; font-weight: 500 !important; }

/* Date input */
.stDateInput > div { background: #fff !important; }

/* Metric */
.stMetric { background: #fff; border: 1.5px solid #e2e8f0; border-radius: 12px; padding: 16px; }
</style>
""", unsafe_allow_html=True)

html = lambda x: st.markdown(x, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# AUTENTICAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
USERS_FILE = os.path.join(_BASE, "usuarios.json")

def _carregar_usuarios() -> dict:
    if not os.path.exists(USERS_FILE):
        salt = secrets.token_hex(16)
        h    = hashlib.sha256((salt + "admin123").encode()).hexdigest()
        users = {"admin": {"senha": h, "salt": salt, "nome": "Administrador", "perfil": "admin", "ativo": True}}
        json.dump(users, open(USERS_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    return json.load(open(USERS_FILE, encoding="utf-8"))

def _verificar_senha(usuario: str, senha: str) -> dict | None:
    users = _carregar_usuarios()
    u = users.get(usuario)
    if not u or not u.get("ativo", True):
        return None
    h = hashlib.sha256((u["salt"] + senha).encode()).hexdigest()
    return u if h == u["senha"] else None

def _salvar_usuarios(users: dict):
    json.dump(users, open(USERS_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
if "user" not in st.session_state:
    st.markdown("""
    <div style="min-height:100vh;background:linear-gradient(135deg,#1e3a5f,#1d4ed8);
                display:flex;align-items:center;justify-content:center">
    </div>""", unsafe_allow_html=True)

    col_c = st.columns([1, 1.2, 1])[1]
    with col_c:
        html("""
        <div style="background:#fff;border-radius:20px;padding:44px 40px;
                    box-shadow:0 20px 60px #1d4ed840;margin-top:60px">
          <div style="text-align:center;margin-bottom:32px">
            <div style="font-size:48px;margin-bottom:8px">⚖️</div>
            <h2 style="margin:0;font-size:22px;font-weight:800;color:#1e3a5f">Controladoria Geral</h2>
            <p style="margin:4px 0 0;color:#64748b;font-size:13px">
              Prefeitura de Governador Edison Lobão/MA</p>
          </div>
        </div>""")

        with st.form("login_form"):
            usuario = st.text_input("Usuário", placeholder="ex: admin")
            senha   = st.text_input("Senha",   type="password", placeholder="••••••••")
            ok = st.form_submit_button("Entrar →", type="primary", use_container_width=True)

        if ok:
            u = _verificar_senha(usuario.strip(), senha)
            if u:
                st.session_state["user"]   = usuario.strip()
                st.session_state["perfil"] = u["perfil"]
                st.session_state["nome"]   = u.get("nome", usuario)
                log.info(f"LOGIN: {usuario}")
                st.rerun()
            else:
                html('<p style="color:#dc2626;font-size:13px;text-align:center;margin-top:8px">❌ Credenciais inválidas</p>')
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# DADOS DO USUÁRIO LOGADO
# ─────────────────────────────────────────────────────────────────────────────
_user_nome   = st.session_state.get("nome", "Usuário")
_user_perfil = st.session_state.get("perfil", "operador")
_is_admin    = _user_perfil == "admin"

# ─────────────────────────────────────────────────────────────────────────────
# CHECKLISTS
# ─────────────────────────────────────────────────────────────────────────────
CHECKLIST_PADRAO = [
    "Validação do Documento Fiscal",
    "Emissão e Autenticação das Certidões Negativas",
    "Conformidade com o processo Licitatório (Preço, Descrição do Produto)",
    "Disponibilidade de Saldos Licitatórios",
    "Outros: Contratos, Valores Somatórios, Impostos, constantes na NF",
    "Extrato do Contrato",
]
CHECKLIST_ENG = [
    "Validação do Documento Fiscal",
    "Emissão e Autenticação das Certidões Negativas",
    "Conformidade com o processo Licitatório (Preço, Descrição do Produto)",
    "Disponibilidade de Saldos Licitatórios",
    "Outros: Contratos, Valores Somatórios, Impostos, constantes na NF",
    "Extrato do Contrato",
    "Solicitação de pagamento da empresa indicando a medição, valor e dados bancários",
    "Planilha de medição assinada pelo engenheiro da empresa e atestada pelo engenheiro da Prefeitura",
    "Relatório fotográfico a partir da 2ª medição atestado pelo engenheiro",
    "Declaração da Fazenda Municipal de origem do registro do Bloco de Notas",
    "Portaria de nomeação do engenheiro da Prefeitura",
    "Cópia Do Contrato",
    "ART ou RRT Do Projeto",
    "ART de Fiscalização",
    "Portaria De Nomeação Do Fiscal",
    "Termo De Recebimento Definitivo Da Obra",
]
CHECKLIST_TDF = [
    "Ofício",
    "Formulário TFD",
    "Conformidade com o processo Licitatório (Preço, Descrição do Produto)",
    "Laudo Médico",
    "Documentos pessoais do beneficiário",
    "Agendamento / Comprovante de consulta",
]
CHECKLIST_PASSAGEM = [
    "Prestação de contas diárias",
    "Documentação comprobatória do deslocamento",
    "Requerimento de restituição",
]
OBS_PASSAGEM = (
    "Em conformidade com o Art. 4º, § 2º, alínea 'a', do Decreto nº 022, de 03 de junho de 2025, "
    "o ressarcimento de passagens aéreas para trechos de até 1.000 km de distância do Município de "
    "Governador Edson Lobão/MA está condicionado a que seus valores não ultrapassem 20% do custo total "
    "(ida e volta) das passagens terrestres para o mesmo destino, Considerando que o valor apresentado "
    "neste requerimento excede o limite estabelecido, esta Controladoria INDEFERE o pedido de restituição "
    "de passagem, ressalvado outro entendimento.\n"
    "Encaminha-se o pedido de requerimento de solicitação à Secretaria de Administração para arquivamento."
)
TIPOS_INFO = {
    "padrao":   {"label": "Anexo II",            "icon": "📄", "cor": "#2563eb", "itens": 6,  "checklist": CHECKLIST_PADRAO},
    "eng":      {"label": "NF Engenharia",        "icon": "🏗️", "cor": "#7c3aed", "itens": 16, "checklist": CHECKLIST_ENG},
    "tdf":      {"label": "TFD",                  "icon": "🏥", "cor": "#0f766e", "itens": 6,  "checklist": CHECKLIST_TDF},
    "passagem": {"label": "Restituição Passagem", "icon": "✈️", "cor": "#d97706", "itens": 3,  "checklist": CHECKLIST_PASSAGEM},
}

# ─────────────────────────────────────────────────────────────────────────────
# UTILITÁRIOS
# ─────────────────────────────────────────────────────────────────────────────
MESES = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
         "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

def data_extenso(d: date) -> str:
    return f"{d.day} de {MESES[d.month]} de {d.year}"

def formatar_moeda(v: str) -> str:
    d = re.sub(r"\D", "", str(v))
    if not d: return ""
    c = int(d); r = c // 100; ct = c % 100
    return f"R$ {r:,}".replace(",", ".") + f",{ct:02d}"

def formatar_cnpj(v: str) -> str:
    d = re.sub(r"\D", "", str(v))[:14]
    if len(d) < 14: return v
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"

def sugerir_tipo(obj: str) -> str:
    o = str(obj).upper()
    if any(x in o for x in ["ENGENHARIA","OBRA","PAVIMENT","CONSTRU","REFORMA","MEDICAO","ART","RRT"]):
        return "eng"
    if any(x in o for x in ["TFD","TRATAMENTO","SAUDE","LAUDO","MÉDICO"]):
        return "tdf"
    if any(x in o for x in ["PASSAGEM","RESTITUICAO","DIARIA","VIAGEM","AÉREO","AEREO"]):
        return "passagem"
    return "padrao"

def valor_fmt(v: str) -> str:
    v = str(v).strip().replace("R$", "").strip()
    try: return f"R$ {float(re.sub(r'[.]','',v.replace(',','.')) if ',' in v else v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    except: return v

def get_ip() -> str:
    try: return socket.gethostbyname(socket.gethostname())
    except: return "127.0.0.1"

def get_proximo_numero() -> int:
    if USAR_GSHEETS: return 0
    return proximo_numero(EXCEL_PATH)

# ─────────────────────────────────────────────────────────────────────────────
# CARREGAR DADOS
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=180)
def load_df() -> pd.DataFrame:
    if USAR_GSHEETS:
        try:
            import gspread
            from google.oauth2.service_account import Credentials
            creds = Credentials.from_service_account_info(
                dict(st.secrets["gcp_service_account"]),
                scopes=["https://www.googleapis.com/auth/spreadsheets.readonly",
                        "https://www.googleapis.com/auth/drive.readonly"])
            gc = gspread.authorize(creds)
            sh = gc.open_by_key(st.secrets["google_sheets"]["sheet_id"])
            data = sh.sheet1.get_all_values()
            df = pd.DataFrame(data[1:], columns=data[0])
        except Exception as e:
            st.error(f"Erro Google Sheets: {e}")
            return pd.DataFrame()
    else:
        if not os.path.exists(EXCEL_PATH):
            return pd.DataFrame()
        df = pd.read_excel(EXCEL_PATH, dtype=str)

    df.columns = [str(c).strip().replace("\xa0", "").replace("\n", "").replace("\t", "") for c in df.columns]
    _aliases = {
        "SECRETARIO": ["SECRETARIO ", "SECRETÁRIO", "SECRETÁRIO "],
        "ORGÃO":      ["ÓRGÃO", "ORGAO", "ÓRGAO"],
        "NÚMERO DO DOCUMENTO": ["NUMERO DO DOCUMENTO", "NÚM DOC"],
    }
    for can, alts in _aliases.items():
        if can not in df.columns:
            for alt in alts:
                m = [c for c in df.columns if c.strip() == alt.strip()]
                if m:
                    df.rename(columns={m[0]: can}, inplace=True)
                    break
    return df.fillna("")

@st.cache_data(ttl=180)
def build_maps(excel_path: str) -> dict:
    """Constrói todos os mapas de autopreenchimento de uma vez."""
    df = load_df()
    if df.empty:
        return {"orgao_sec": {}, "forn_cnpj": {}, "mod_cont": {}, "cnpj_forn": {}}
    orgao_sec = {}; forn_cnpj = {}; mod_cont = {}; cnpj_forn = {}
    for _, r in df.iterrows():
        org  = str(r.get("ORGÃO", "")).strip()
        sec  = str(r.get("SECRETARIO", "")).strip()
        forn = str(r.get("FORNECEDOR", "")).strip()
        cnpj = str(r.get("CNPJ", "")).strip()
        mod  = str(r.get("MODALIDADE", "")).strip()
        cont = str(r.get("CONTRATO", "")).strip()
        if org  and sec:  orgao_sec[org]  = sec
        if forn and cnpj: forn_cnpj[forn] = cnpj
        if mod  and cont: mod_cont[mod]   = cont
        if cnpj and forn: cnpj_forn[cnpj] = forn
    return {"orgao_sec": orgao_sec, "forn_cnpj": forn_cnpj,
            "mod_cont": mod_cont, "cnpj_forn": cnpj_forn}

def salvar_historico(row: dict):
    if USAR_GSHEETS or not os.path.exists(EXCEL_PATH):
        return
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Histórico" not in wb.sheetnames:
            ws = wb.create_sheet("Histórico")
            ws.append(["Data/Hora", "Processo", "Órgão", "Fornecedor",
                        "Valor", "Tipo Parecer", "Decisão", "Operador"])
        else:
            ws = wb["Histórico"]
        ws.append([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            row.get("processo", ""), row.get("orgao", ""),
            row.get("fornecedor", ""), row.get("valor", ""),
            row.get("tipo_label", ""), row.get("decisao", ""),
            row.get("operador", _user_nome),
        ])
        wb.save(EXCEL_PATH)
        st.cache_data.clear()
    except Exception as e:
        log.warning(f"Histórico: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    html(f"""
    <div style="padding:24px 20px 16px">
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:4px">
        <div style="font-size:28px">⚖️</div>
        <div>
          <div style="font-weight:800;font-size:14px;color:#fff">Controladoria</div>
          <div style="font-size:10px;color:#93c5fd">Gov. Edison Lobão/MA</div>
        </div>
      </div>
    </div>
    <div style="height:1px;background:#ffffff18;margin:0 20px 16px"></div>
    <div style="padding:0 12px">
      <div style="background:#ffffff18;border-radius:10px;padding:10px 14px;margin-bottom:16px">
        <div style="font-size:10px;color:#93c5fd;font-weight:600;text-transform:uppercase;letter-spacing:.07em">Logado como</div>
        <div style="font-weight:700;font-size:13px;color:#fff;margin-top:2px">{_user_nome}</div>
        <div style="font-size:10px;color:#93c5fd;margin-top:1px;text-transform:capitalize">{_user_perfil}</div>
      </div>
    </div>
    """)

    pagina = st.radio("Navegação", [
        "🏠  Dashboard",
        "⚡  Gerar Documentos",
        "➕  Novo Processo",
        "🔍  Buscar / Editar",
        "📊  Histórico",
        "👥  Usuários",
        "⚙️  Configurações",
    ], label_visibility="collapsed")

    # Status da planilha
    html('<div style="height:1px;background:#ffffff18;margin:16px 20px"></div>')
    n_prox = get_proximo_numero()
    modo = "☁️ Google Sheets" if USAR_GSHEETS else "💾 Excel Local"
    html(f"""
    <div style="padding:0 20px 8px">
      <div style="font-size:10px;color:#93c5fd;font-weight:600;text-transform:uppercase">Próximo Nº</div>
      <div style="font-size:26px;font-weight:800;color:#60a5fa;margin:2px 0">{n_prox or '—'}</div>
      <div style="font-size:10px;color:#7dd3fc">{modo}</div>
    </div>
    """)
    if st.button("🔄 Atualizar Dados", use_container_width=True, key="sb_refresh"):
        st.cache_data.clear()
        st.rerun()

    # Logout
    html('<div style="height:1px;background:#ffffff18;margin:12px 20px"></div>')
    if st.button("🚪 Sair", use_container_width=True, key="sb_logout"):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    # IP
    ip = get_ip()
    html(f'<div style="padding:8px 20px 16px;font-size:10px;color:#475569">{ip}:8501 · v5.0</div>')

# ─────────────────────────────────────────────────────────────────────────────
# HELPER: PAGE HEADER
# ─────────────────────────────────────────────────────────────────────────────
def page_header(icon: str, title: str, subtitle: str = "", cor: str = "#1d4ed8"):
    html(f"""
    <div style="background:linear-gradient(135deg,#1e3a5f,{cor});
                padding:22px 32px;margin-bottom:20px">
      <div style="display:flex;align-items:center;gap:14px">
        <div style="font-size:28px">{icon}</div>
        <div>
          <h2 style="color:#fff;margin:0;font-size:20px;font-weight:800">{title}</h2>
          {f'<p style="color:#93c5fd;margin:3px 0 0;font-size:13px">{subtitle}</p>' if subtitle else ''}
        </div>
      </div>
    </div>""")

# ─────────────────────────────────────────────────────────────────────────────
# ══ PÁGINA: DASHBOARD ════════════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
if pagina == "🏠  Dashboard":
    page_header("🏠", "Dashboard", "Visão geral em tempo real do sistema")

    with st.spinner("Carregando dados..."):
        df = load_df()

    pad = {"padding": "0 32px 32px"}
    html('<div style="padding:0 32px">')

    if df.empty:
        st.warning("⚠️ Planilha não encontrada. Configure o caminho em ⚙️ Configurações.")
        html('</div>')
        st.stop()

    # ── KPIs ──────────────────────────────────────────────────────────────────
    total       = len(df)
    n_orgaos    = df["ORGÃO"].nunique()          if "ORGÃO"      in df.columns else 0
    n_forn      = df["FORNECEDOR"].nunique()     if "FORNECEDOR" in df.columns else 0
    n_este_mes  = 0

    # Tenta calcular processos do mês atual com coluna DATA
    if "DATA" in df.columns:
        mes_atual = datetime.now().month
        ano_atual = datetime.now().year
        def mes_match(v):
            try:
                s = str(v).strip()
                # formato "dd/mm/aaaa" ou extenso
                if "/" in s:
                    p = s.split("/")
                    return int(p[1]) == mes_atual and int(p[2]) == ano_atual
                if str(ano_atual) in s and MESES[mes_atual] in s.lower():
                    return True
            except: pass
            return False
        n_este_mes = df["DATA"].apply(mes_match).sum()

    k1, k2, k3, k4 = st.columns(4)
    for col, val, lbl, delta, cor in [
        (k1, total,      "Total de Processos",      f"+{n_este_mes} este mês", "#2563eb"),
        (k2, n_orgaos,   "Órgãos Distintos",        "na planilha",             "#7c3aed"),
        (k3, n_forn,     "Fornecedores Únicos",     "cadastrados",             "#0f766e"),
        (k4, n_este_mes, "Processos Este Mês",       "período atual",           "#d97706"),
    ]:
        col.markdown(f"""
        <div class="kpi">
          <div class="kpi-lbl">{lbl}</div>
          <div class="kpi-val" style="color:{cor}">{val:,}</div>
          <div class="kpi-delta up">{delta}</div>
        </div>""", unsafe_allow_html=True)

    html('<div style="height:20px"></div>')

    # ── GRÁFICOS ──────────────────────────────────────────────────────────────
    try:
        import plotly.express as px
        col_g1, col_g2 = st.columns([3, 2])

        with col_g1:
            html('<div class="card">')
            html('<p class="sec-hdr">📊 Processos por Órgão (Top 10)</p>')
            if "ORGÃO" in df.columns:
                top_org = df["ORGÃO"].value_counts().head(10).reset_index()
                top_org.columns = ["Órgão", "Processos"]
                fig = px.bar(top_org, x="Processos", y="Órgão", orientation="h",
                             color="Processos", color_continuous_scale="Blues",
                             template="plotly_white", height=300)
                fig.update_layout(margin=dict(l=0,r=0,t=10,b=0), showlegend=False,
                                  coloraxis_showscale=False, yaxis_title="",
                                  plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig, use_container_width=True)
            html('</div>')

        with col_g2:
            html('<div class="card">')
            html('<p class="sec-hdr">🏷️ Distribuição por Modalidade</p>')
            if "MODALIDADE" in df.columns:
                mods = df["MODALIDADE"].value_counts().head(8)
                fig2 = px.pie(values=mods.values, names=mods.index,
                              color_discrete_sequence=px.colors.sequential.Blues_r,
                              template="plotly_white", height=300)
                fig2.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                                   legend=dict(font=dict(size=10)),
                                   plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig2, use_container_width=True)
            html('</div>')

    except ImportError:
        st.info("💡 Instale `plotly` para ver gráficos: `pip install plotly`")

    # ── ÚLTIMOS PROCESSOS ─────────────────────────────────────────────────────
    col_r1, col_r2 = st.columns([2, 1])
    with col_r1:
        html('<div class="card">')
        html('<p class="sec-hdr">🕐 Últimos 10 Processos Cadastrados</p>')
        cols_show = [c for c in ["NÚMERO DO DOCUMENTO","ORGÃO","FORNECEDOR","VALOR","DATA"]
                     if c in df.columns]
        st.dataframe(df[cols_show].tail(10).iloc[::-1].reset_index(drop=True),
                     use_container_width=True, hide_index=True,
                     column_config={
                         "NÚMERO DO DOCUMENTO": st.column_config.TextColumn("Nº", width="small"),
                         "VALOR":               st.column_config.TextColumn("Valor"),
                     })
        html('</div>')

    with col_r2:
        html('<div class="card">')
        html('<p class="sec-hdr">📈 Estatísticas Rápidas</p>')
        if "VALOR" in df.columns:
            def parse_valor(v):
                try: return float(re.sub(r"[R$\s.]", "", str(v)).replace(",", "."))
                except: return 0.0
            vals = df["VALOR"].apply(parse_valor)
            total_v = vals.sum(); media_v = vals[vals > 0].mean()
            st.metric("Valor Total", f"R$ {total_v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            st.metric("Média por Processo", f"R$ {media_v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if media_v else "—")
        st.metric("Maior CNPJ repetido",
                  df["CNPJ"].value_counts().index[0] if "CNPJ" in df.columns and len(df) > 0 else "—")
        html('</div>')

    html('</div>')

# ─────────────────────────────────────────────────────────────────────────────
# ══ PÁGINA: GERAR DOCUMENTOS ════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "⚡  Gerar Documentos":

    # STATE
    def s(k, v=None):
        if k not in st.session_state: st.session_state[k] = v
        return st.session_state[k]

    s("etapa", 1); s("proc_sel", None); s("tipo", "padrao")
    s("decisao", "deferir"); s("b_capa", None); s("b_parecer", None)

    etapa = st.session_state["etapa"]
    proc  = st.session_state["proc_sel"]

    def ir(e): st.session_state["etapa"] = e

    # Barra de progresso visual
    steps = [("1", "Buscar"), ("2", "Preencher"), ("3", "Gerar")]
    step_html = '<div style="display:flex;align-items:center;background:#fff;border-bottom:2px solid #f1f5f9;padding:0 32px">'
    for i, (n, lbl) in enumerate(steps):
        if etapa > int(n):
            dc, lc = "background:#059669;color:#fff", "color:#059669;font-weight:700"
            sym = "✓"
        elif etapa == int(n):
            dc, lc = "background:#1d4ed8;color:#fff", "color:#1d4ed8;font-weight:700"
            sym = n
        else:
            dc, lc = "background:#f1f5f9;color:#94a3b8", "color:#94a3b8;font-weight:500"
            sym = n
        step_html += f"""
        <div style="display:flex;align-items:center;gap:10px;padding:14px 0">
          <div style="width:28px;height:28px;border-radius:50%;display:flex;align-items:center;
                      justify-content:center;font-size:12px;font-weight:800;{dc}">{sym}</div>
          <span style="font-size:13px;{lc}">{lbl}</span>
        </div>"""
        if i < 2:
            div_cor = "#059669" if etapa > int(n) else "#e2e8f0"
            step_html += f'<div style="flex:1;height:2px;background:{div_cor};margin:0 12px;border-radius:1px"></div>'
    step_html += "</div>"
    html(step_html)

    html('<div style="padding:0 32px 32px;margin-top:20px">')

    # ── ETAPA 1: BUSCA INTELIGENTE ────────────────────────────────────────────
    if etapa == 1:
        html('<h3 style="font-size:18px;font-weight:800;color:#0f172a;margin:0 0 4px">Buscar Processo</h3>')
        html('<p style="color:#64748b;font-size:14px;margin:0 0 20px">Use os filtros ou selecione diretamente — processos recentes aparecem automaticamente</p>')

        with st.spinner("Carregando..."):
            df = load_df()

        if df.empty:
            st.error("❌ Planilha não encontrada.")
            html('</div>')
            st.stop()

        # Filtros inteligentes
        html('<div class="card" style="padding:18px 20px">')
        fc1, fc2, fc3, fc4, fc5 = st.columns([1.5, 2, 2, 1.5, 0.7])
        q_num    = fc1.text_input("Nº Documento",    placeholder="ex: 2465")
        forns    = ["Todos"] + sorted(df["FORNECEDOR"].dropna().unique().tolist()) if "FORNECEDOR" in df.columns else ["Todos"]
        orgs     = ["Todos"] + sorted(df["ORGÃO"].dropna().unique().tolist())      if "ORGÃO"      in df.columns else ["Todos"]
        forn_sel = fc2.selectbox("Fornecedor", forns,   index=0)
        org_sel  = fc3.selectbox("Órgão",      orgs,    index=0)
        tipo_sel = fc4.selectbox("Tipo", ["Todos", "Padrão", "Engenharia", "TDF", "Passagem"], index=0)
        with fc5:
            st.write("")
            if st.button("🔄", help="Atualizar", use_container_width=True):
                st.cache_data.clear(); st.rerun()
        html('</div>')

        # Aplica filtros
        dff = df.copy()
        if q_num:
            dff = dff[dff["NÚMERO DO DOCUMENTO"].astype(str).str.contains(q_num, case=False, na=False)]
        if forn_sel != "Todos" and "FORNECEDOR" in dff.columns:
            dff = dff[dff["FORNECEDOR"] == forn_sel]
        if org_sel != "Todos" and "ORGÃO" in dff.columns:
            dff = dff[dff["ORGÃO"] == org_sel]
        if tipo_sel != "Todos":
            mapa = {"Padrão": "padrao", "Engenharia": "eng", "TDF": "tdf", "Passagem": "passagem"}
            t_f = mapa[tipo_sel]
            dff = dff[dff["OBJETO"].apply(lambda x: sugerir_tipo(x) == t_f)] if "OBJETO" in dff.columns else dff

        # Se não há filtros digitados, mostra os 20 mais recentes
        if not q_num and forn_sel == "Todos" and org_sel == "Todos" and tipo_sel == "Todos":
            html("""<div style="display:flex;align-items:center;gap:8px;margin:12px 0 6px">
              <span style="font-size:11px;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:.07em">
              ⏱️ Processos Recentes (20 últimos)</span>
            </div>""")
            dff_show = dff.tail(20).iloc[::-1]
        else:
            html(f'<div style="font-size:12px;color:#64748b;margin:8px 0 4px"><b>{len(dff)}</b> resultado(s) de <b>{len(df)}</b></div>')
            dff_show = dff.head(100)

        cols_e = [c for c in ["NÚMERO DO DOCUMENTO", "ORGÃO", "FORNECEDOR", "VALOR", "DATA", "MODALIDADE"]
                  if c in dff_show.columns]
        if not dff_show.empty:
            st.dataframe(dff_show[cols_e], use_container_width=True, hide_index=True, height=260,
                         column_config={
                             "NÚMERO DO DOCUMENTO": st.column_config.TextColumn("Nº Doc.", width="small"),
                             "VALOR":               st.column_config.TextColumn("Valor"),
                             "ORGÃO":               st.column_config.TextColumn("Órgão"),
                         })
            html('<p class="sec-hdr">Selecione o processo</p>')
            opcoes = ["— selecione —"] + dff_show["NÚMERO DO DOCUMENTO"].unique().tolist()
            num_sel = st.selectbox("", opcoes, label_visibility="collapsed")
            if num_sel != "— selecione —":
                ln = dff_show[dff_show["NÚMERO DO DOCUMENTO"] == num_sel].iloc[0]
                st.session_state["proc_sel"] = ln.to_dict()
                st.session_state["tipo"]     = sugerir_tipo(ln.get("OBJETO", ""))
                ir(2); st.rerun()
        else:
            st.warning("⚠️ Nenhum processo encontrado.")

    # ── ETAPA 2: FORMULÁRIO ───────────────────────────────────────────────────
    elif etapa == 2 and proc:
        num_proc = proc.get("NÚMERO DO DOCUMENTO", "")
        tipo     = st.session_state["tipo"]
        ti       = TIPOS_INFO[tipo]

        # Banner de autopreenchimento
        html(f"""
        <div style="background:linear-gradient(135deg,#eff6ff,#f0fdf4);
                    border:1.5px solid #bfdbfe;border-radius:12px;padding:12px 18px;
                    display:flex;align-items:center;gap:12px;margin-bottom:18px">
          <span style="font-size:22px">✨</span>
          <div>
            <span style="font-weight:700;color:#1d4ed8;font-size:14px">Autopreenchimento aplicado!</span>
            <span style="color:#475569;font-size:13px"> Processo <b>#{num_proc}</b> — edite se necessário.</span>
          </div>
          <div style="margin-left:auto">
            <span style="background:#dbeafe;color:#1d4ed8;padding:3px 10px;border-radius:20px;
                         font-size:11px;font-weight:700">🔒 Dados da planilha</span>
          </div>
        </div>""")

        # Seletor de tipo
        html('<p class="sec-hdr">Tipo de Parecer</p>')
        t_cols = st.columns(4)
        for col, (tid, tinfo) in zip(t_cols, TIPOS_INFO.items()):
            ativo = st.session_state["tipo"] == tid
            with col:
                brd = tinfo["cor"] if ativo else "#e2e8f0"
                bg  = tinfo["cor"] + "18" if ativo else "#fff"
                html(f"""
                <div style="border:2px solid {brd};background:{bg};border-radius:12px;
                            padding:12px 8px;text-align:center;margin-bottom:6px">
                  <div style="font-size:20px">{tinfo['icon']}</div>
                  <div style="font-size:12px;font-weight:700;color:{''+tinfo['cor'] if ativo else '#374151'}">{tinfo['label']}</div>
                  <div style="font-size:10px;color:#94a3b8">{tinfo['itens']} itens</div>
                </div>""")
                if st.button("✓ Ativo" if ativo else "Selecionar", key=f"t_{tid}",
                             use_container_width=True, type="primary" if ativo else "secondary"):
                    st.session_state["tipo"] = tid; st.rerun()

        tipo = st.session_state["tipo"]; ti = TIPOS_INFO[tipo]
        st.divider()

        # Carrega mapas
        maps = build_maps(EXCEL_PATH)
        df_ac = load_df()

        col_capa, col_parecer = st.columns(2, gap="large")

        with col_capa:
            html("""<div style="background:linear-gradient(135deg,#1e3a5f,#1d4ed8);
                                border-radius:12px 12px 0 0;padding:12px 16px;
                                display:flex;align-items:center;gap:8px">
              <span style="font-size:16px">📋</span>
              <span style="color:#fff;font-weight:700;font-size:14px">Capa do Processo</span>
              <span style="margin-left:auto;font-size:10px;background:#ffffff25;color:#bfdbfe;
                           padding:3px 8px;border-radius:5px">Comum a todos os tipos</span>
            </div>
            <div style="border:1.5px solid #e2e8f0;border-top:none;border-radius:0 0 12px 12px;
                        padding:18px;background:#fff">""")

            processo    = st.text_input("**Nº do Processo**",       value=proc.get("NÚMERO DO DOCUMENTO",""), key="fp")
            orgao       = st.text_input("**Órgão / Secretaria**",   value=proc.get("ORGÃO",""),              key="fo")

            # Secretário: auto por órgão
            _org_val = proc.get("ORGÃO","").strip()
            _sec_auto = maps["orgao_sec"].get(_org_val, proc.get("SECRETARIO","").strip())
            if _sec_auto != proc.get("SECRETARIO","").strip():
                html(f'<p class="field-auto">🔗 Auto: {_sec_auto}</p>')

            fornecedor  = st.text_input("**Fornecedor**",           value=proc.get("FORNECEDOR",""),          key="ff")

            ca, cb = st.columns(2)
            # CNPJ: auto por fornecedor
            _forn_val = proc.get("FORNECEDOR","").strip()
            _cnpj_auto = maps["forn_cnpj"].get(_forn_val, proc.get("CNPJ",""))
            cnpj       = ca.text_input("**CNPJ**", value=_cnpj_auto, key="fc")
            nf_capa    = cb.text_input("**NF/Fatura**", value=str(proc.get("Nº","")), key="fn")

            ca2, cb2 = st.columns(2)
            contrato   = ca2.text_input("**Contrato**",   value=proc.get("CONTRATO",""),   key="fct")
            modalidade = cb2.text_input("**Modalidade**", value=proc.get("MODALIDADE",""), key="fm")

            ca3, cb3 = st.columns(2)
            periodo_ref  = ca3.text_input("**Período Ref.**",    value=proc.get("PERÍODO DE REFERÊNCIA",""), key="fpr")
            ordem_compra = cb3.text_input("**N° Ordem Compra**", value=proc.get("N° ORDEM DE COMPRA",""),   key="foc", placeholder="Opcional")

            ca4, cb4 = st.columns(2)
            data_nf    = ca4.text_input("**Data da NF**",     value=proc.get("DATA NF",""),  key="fdn")
            secretario = cb4.text_input("**Secretário(a)**",  value=_sec_auto,               key="fs")

            data_ateste_d   = st.date_input("**Data do Ateste**", value=date.today(), key="fda", format="DD/MM/YYYY")
            data_ateste_str = data_extenso(data_ateste_d)
            html(f'<p style="font-size:11px;color:#64748b;margin:2px 0 0">📅 {data_ateste_str}</p>')
            html('</div>')

        with col_parecer:
            html(f"""<div style="background:{ti['cor']};border-radius:12px 12px 0 0;
                                padding:12px 16px;display:flex;align-items:center;gap:8px">
              <span style="font-size:16px">{ti['icon']}</span>
              <span style="color:#fff;font-weight:700;font-size:14px">Parecer — {ti['label']}</span>
              <span style="margin-left:auto;font-size:10px;background:#ffffff25;color:#fff;
                           padding:3px 8px;border-radius:5px">{ti['itens']} itens</span>
            </div>
            <div style="border:1.5px solid {ti['cor']}40;border-top:none;border-radius:0 0 12px 12px;
                        padding:18px;background:#fff">""")

            if tipo == "passagem":
                objeto_p    = st.text_area("**Objeto**", value=proc.get("OBJETO",""), height=68, key="op")
                doc_oficio  = st.text_input("**Documento (Ofício)**", value="", key="dof", placeholder="OFÍCIO N° 136.2025-SEMAD")
                cp1, cp2 = st.columns(2)
                solicitante = cp1.text_input("**Solicitante**", value=_sec_auto, key="sol")
                cpf_sol     = cp2.text_input("**CPF**",         value="", key="cpf", placeholder="000.000.000-00")
                valor_p     = st.text_input("**Valor**",         value=proc.get("VALOR",""), key="vp")
            else:
                objeto_n    = st.text_area("**Objeto da Despesa**",  value=proc.get("OBJETO",""),           height=68, key="on")
                cn1, cn2 = st.columns(2)
                tipo_doc_n  = cn1.text_input("**Tipo Doc. Fiscal**", value=proc.get("DOCUMENTO FISCAL",""), key="tdn")
                nf_n        = cn2.text_input("**Nº da NF**",         value=str(proc.get("Nº","")),          key="nfn")
                cn3, cn4 = st.columns(2)
                tipo_nf_n   = cn3.text_input("**Tipo NF**",          value=proc.get("TIPO",""),             key="tnfn")
                valor_n     = cn4.text_input("**Valor**",             value=proc.get("VALOR",""),            key="vn")
                fornec_n    = st.text_input("**Fornecedor (Parecer)**", value=proc.get("FORNECEDOR",""),     key="pfn")
                cnpj_n      = st.text_input("**CNPJ (Parecer)**",      value=_cnpj_auto,                    key="pcn")

            # Decisão
            html('<p class="sec-hdr">Decisão</p>')
            dc1, dc2 = st.columns(2)
            dec = st.session_state.get("decisao", "deferir")
            with dc1:
                if st.button("✅  Deferir",    key="bd", use_container_width=True, type="primary" if dec=="deferir" else "secondary"):
                    st.session_state["decisao"] = "deferir"; st.rerun()
            with dc2:
                if st.button("❌  Indeferir", key="bi", use_container_width=True, type="primary" if dec=="indeferir" else "secondary"):
                    st.session_state["decisao"] = "indeferir"; st.rerun()

            cor_dec = "#059669" if dec=="deferir" else "#dc2626"
            html(f'<p style="font-size:13px;font-weight:700;color:{cor_dec};margin:6px 0 10px">{"✅ DEFERIMOS O PAGAMENTO" if dec=="deferir" else "❌ INDEFERIMOS O PAGAMENTO"}</p>')

            obs_default = OBS_PASSAGEM if tipo == "passagem" else ""
            obs = st.text_area("**Observação**", value=obs_default, height=90, key="obs",
                               placeholder="Deixe em branco se não houver...")
            html('</div>')

            # Checklist moderno
            chk_key = f"chk_{tipo}"
            n_itens = ti["itens"]
            if chk_key not in st.session_state: st.session_state[chk_key] = [True] * n_itens
            if len(st.session_state[chk_key]) != n_itens: st.session_state[chk_key] = [True] * n_itens

            qtd_ok  = sum(st.session_state[chk_key])
            qtd_err = n_itens - qtd_ok
            pct     = int(qtd_ok / n_itens * 100) if n_itens else 100
            bar_cor = "#16a34a" if pct == 100 else "#f59e0b" if pct >= 60 else "#dc2626"

            html(f"""
            <div style="background:#fff;border:1.5px solid {ti['cor']}30;border-radius:12px;
                        padding:14px 16px;margin-top:12px">
              <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
                <div>
                  <div style="font-weight:700;font-size:13px;color:#0f172a">☑ Checklist — {ti['label']}</div>
                  <div style="font-size:11px;color:#94a3b8;margin-top:2px">
                    <span style="color:#16a34a;font-weight:600">{qtd_ok} corretos</span>
                    {f' · <span style="color:#dc2626;font-weight:600">{qtd_err} pendentes</span>' if qtd_err else ''}
                  </div>
                </div>
                <div style="background:#f1f5f9;border-radius:6px;padding:2px;height:8px;
                            width:80px;overflow:hidden">
                  <div style="height:100%;width:{pct}%;background:{bar_cor};border-radius:4px;transition:width .3s"></div>
                </div>
              </div>""")

            ga1, ga2 = st.columns(2)
            with ga1:
                if st.button("✓ Todos Corretos", key=f"all_ok_{tipo}", use_container_width=True):
                    st.session_state[chk_key] = [True]  * n_itens; st.rerun()
            with ga2:
                if st.button("✗ Todos Errados",  key=f"all_err_{tipo}", use_container_width=True):
                    st.session_state[chk_key] = [False] * n_itens; st.rerun()

            html('<div style="margin-top:8px">')
            for i, item in enumerate(ti["checklist"]):
                ok    = st.session_state[chk_key][i]
                bg_c  = "#f0fdf4" if ok else "#fff5f5"
                brd_c = "#bbf7d0" if ok else "#fecaca"
                sym   = "✓" if ok else "✗"
                sym_c = "#16a34a" if ok else "#dc2626"
                html(f"""<div style="display:flex;align-items:center;gap:8px;padding:6px 10px;
                              border-radius:8px;margin-bottom:4px;background:{bg_c};
                              border:1px solid {brd_c}">
                  <span style="font-size:15px;font-weight:800;color:{sym_c};min-width:18px;text-align:center">{sym}</span>
                  <span style="font-size:10px;font-weight:700;color:#94a3b8;min-width:16px">{i+1}.</span>
                  <span style="font-size:12px;color:#374151;flex:1">{item}</span>
                </div>""")
                if st.button("✗ Errado" if ok else "✓ Correto",
                             key=f"chk_{tipo}_{i}", use_container_width=True):
                    st.session_state[chk_key][i] = not st.session_state[chk_key][i]; st.rerun()
            html('</div></div>')

        # ── BOTÕES DE AÇÃO ────────────────────────────────────────────────────
        html('<div style="margin-top:24px"></div>')
        st.divider()
        b1, b2, b3, b4 = st.columns([2, 2, 2, 1])

        deferir  = st.session_state.get("decisao", "deferir") == "deferir"
        chk_sits = st.session_state.get(chk_key, [True] * n_itens)

        d_comum = {
            "processo":    processo,    "orgao":       orgao,
            "secretario":  secretario,  "fornecedor":  fornecedor,
            "cnpj":        cnpj,        "contrato":    contrato,
            "modalidade":  modalidade,  "nf":          nf_capa,
            "data_nf":     data_nf,     "periodo_ref": periodo_ref,
            "ordem_compra":ordem_compra,"data_ateste": data_ateste_str,
            "obs":         st.session_state.get("obs",""),
        }

        # Gerar Capa
        with b1:
            if st.button("📋  Gerar Capa", use_container_width=True, type="primary", key="btn_capa"):
                with st.spinner("Gerando capa..."):
                    try:
                        log.info(f"Gerando CAPA processo={processo} orgao={orgao}")
                        b = gerar_pdf_capa({**d_comum, "obs_capa": d_comum["obs"]})
                        st.session_state["b_capa"] = b
                        log.info(f"Capa gerada: {len(b)} bytes")
                    except Exception as e:
                        log.error(f"Erro capa: {e}")
                        st.error(f"❌ Erro: {e}")

        if st.session_state.get("b_capa"):
            b1.download_button("⬇️ Baixar Capa", st.session_state["b_capa"],
                                file_name=f"CAPA_{processo}.pdf",
                                mime="application/pdf", use_container_width=True)

        # Gerar Parecer
        with b2:
            if st.button(f"{ti['icon']}  Gerar Parecer", use_container_width=True, type="primary", key="btn_par"):
                with st.spinner(f"Gerando parecer {ti['label']}..."):
                    try:
                        log.info(f"Gerando PARECER tipo={tipo} processo={processo}")
                        if tipo == "passagem":
                            d_pass = {**d_comum,
                                      "objeto":     st.session_state.get("op",""),
                                      "documento":  st.session_state.get("dof",""),
                                      "solicitante":st.session_state.get("sol",""),
                                      "cpf":        st.session_state.get("cpf",""),
                                      "valor":      st.session_state.get("vp","")}
                            b = gerar_pdf_parecer_passagem(d_pass, deferir, ti["checklist"], chk_sits)
                        else:
                            d_par = {**d_comum,
                                     "objeto":   st.session_state.get("on",""),
                                     "tipo_doc": st.session_state.get("tdn",""),
                                     "nf":       st.session_state.get("nfn",""),
                                     "tipo_nf":  st.session_state.get("tnfn",""),
                                     "valor":    st.session_state.get("vn","")}
                            b = gerar_pdf_parecer_padrao(d_par, tipo, deferir, ti["checklist"], chk_sits)
                        st.session_state["b_parecer"] = b
                        log.info(f"Parecer gerado: {len(b)} bytes")
                        salvar_historico({**d_comum, "tipo_label": ti["label"],
                                          "decisao": "DEFERIDO" if deferir else "INDEFERIDO"})
                    except Exception as e:
                        log.error(f"Erro parecer: {e}")
                        st.error(f"❌ Erro ao gerar parecer: {e}")

        if st.session_state.get("b_parecer"):
            b2.download_button("⬇️ Baixar Parecer", st.session_state["b_parecer"],
                                file_name=f"PARECER_{ti['label'].replace(' ','_')}_{processo}.pdf",
                                mime="application/pdf", use_container_width=True)

        # Ambos
        with b3:
            if st.button("📦  Gerar Ambos", use_container_width=True, key="btn_ambos"):
                with st.spinner("Gerando documentos..."):
                    erros_log = []
                    try:
                        log.info(f"Gerando AMBOS processo={processo}")
                        bc = gerar_pdf_capa({**d_comum, "obs_capa": d_comum["obs"]})
                        st.session_state["b_capa"] = bc
                        log.info(f"  Capa: {len(bc)} bytes ✓")
                    except Exception as e:
                        erros_log.append(f"Capa: {e}"); log.error(f"Erro capa: {e}")
                    try:
                        if tipo == "passagem":
                            d_pass = {**d_comum, "objeto": st.session_state.get("op",""),
                                      "documento":st.session_state.get("dof",""),
                                      "solicitante":st.session_state.get("sol",""),
                                      "cpf":st.session_state.get("cpf",""),
                                      "valor":st.session_state.get("vp","")}
                            bp = gerar_pdf_parecer_passagem(d_pass, deferir, ti["checklist"], chk_sits)
                        else:
                            d_par = {**d_comum, "objeto":st.session_state.get("on",""),
                                     "tipo_doc":st.session_state.get("tdn",""),
                                     "nf":st.session_state.get("nfn",""),
                                     "tipo_nf":st.session_state.get("tnfn",""),
                                     "valor":st.session_state.get("vn","")}
                            bp = gerar_pdf_parecer_padrao(d_par, tipo, deferir, ti["checklist"], chk_sits)
                        st.session_state["b_parecer"] = bp
                        log.info(f"  Parecer: {len(bp)} bytes ✓")
                        salvar_historico({**d_comum, "tipo_label": ti["label"],
                                          "decisao": "DEFERIDO" if deferir else "INDEFERIDO"})
                    except Exception as e:
                        erros_log.append(f"Parecer: {e}"); log.error(f"Erro parecer: {e}")

                    if erros_log:
                        for e in erros_log: st.error(f"❌ {e}")
                    else:
                        st.success("✅ Ambos os documentos gerados!")

        with b4:
            if st.button("↩️  Voltar", use_container_width=True, key="btn_voltar"):
                ir(1); st.session_state["proc_sel"] = None; st.rerun()

    html('</div>')  # fecha padding

# ─────────────────────────────────────────────────────────────────────────────
# ══ PÁGINA: NOVO PROCESSO ════════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "➕  Novo Processo":
    page_header("➕", "Cadastrar Novo Processo",
                "Preencha os dados para incluir na planilha", "#7c3aed")

    if not USAR_GSHEETS and not os.path.exists(EXCEL_PATH):
        st.error(f"❌ Planilha não encontrada: `{EXCEL_PATH}`")
        st.stop()

    def salvar_gsheets_linha(nova_linha: dict):
        import gspread
        from google.oauth2.service_account import Credentials
        creds = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]),
            scopes=["https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive"])
        gc = gspread.authorize(creds)
        ws = gc.open_by_key(st.secrets["google_sheets"]["sheet_id"]).sheet1
        headers = ws.row_values(1) or list(nova_linha.keys())
        ws.append_row([nova_linha.get(h, "") for h in headers], value_input_option="USER_ENTERED")

    df_ac = load_df()
    maps  = build_maps(EXCEL_PATH)

    lista_forn  = sorted(df_ac["FORNECEDOR"].dropna().unique().tolist()) if not df_ac.empty and "FORNECEDOR" in df_ac.columns else []
    lista_orgaos= sorted(df_ac["ORGÃO"].dropna().unique().tolist())      if not df_ac.empty and "ORGÃO"      in df_ac.columns else []
    lista_mods  = sorted(df_ac["MODALIDADE"].dropna().unique().tolist()) if not df_ac.empty and "MODALIDADE" in df_ac.columns else []
    lista_objs  = sorted(df_ac["OBJETO"].dropna().unique().tolist())     if not df_ac.empty and "OBJETO"     in df_ac.columns else []

    def si(k, v=""):
        if k not in st.session_state: st.session_state[k] = v

    _prox = get_proximo_numero()
    si("np_saved", False); si("np_saved_data", {})

    if st.session_state.get("np_saved"):
        sd = st.session_state.get("np_saved_data", {})
        html(f"""
        <div style="margin:0 32px 24px;background:linear-gradient(135deg,#f0fdf4,#dcfce7);
                    border:1.5px solid #86efac;border-radius:14px;padding:20px 24px;
                    display:flex;align-items:flex-start;gap:16px">
          <div style="width:44px;height:44px;border-radius:50%;background:#059669;
               display:flex;align-items:center;justify-content:center;font-size:22px;
               flex-shrink:0;color:#fff">✓</div>
          <div>
            <div style="font-weight:800;font-size:16px;color:#15803d;margin-bottom:6px">
              Processo #{sd.get('num','')} cadastrado!</div>
            <div style="font-size:13px;color:#166534">
              <b>Fornecedor:</b> {sd.get('forn','')} &nbsp;·&nbsp;
              <b>Valor:</b> {sd.get('valor','')}
            </div>
          </div>
        </div>""")
        cn, _ = st.columns([1, 3])
        with cn:
            if st.button("➕ Cadastrar Outro", type="primary", use_container_width=True):
                for k in ["np_saved","np_saved_data"]: st.session_state.pop(k, None)
                st.rerun()
        st.stop()

    html('<div style="padding:0 32px 32px">')

    # ── Identificação
    html('<div class="card">')
    html('<p class="sec-hdr">🔢 Identificação</p>')
    ni1, ni2, ni3 = st.columns([1.2, 2, 2])
    n_num_doc  = ni1.text_input("Nº do Documento *", placeholder=str(_prox) if _prox else "ex: 3096", key="np_num_inp")
    if _prox: ni1.caption(f"🔢 Sugerido: {_prox}")
    n_data     = ni2.date_input("Data do Processo *", value=date.today(), key="np_data",    format="DD/MM/YYYY")
    n_data_pub = ni3.date_input("Data de Publicação", value=date.today(), key="np_datapub", format="DD/MM/YYYY")
    html('</div>')

    # ── Fornecedor
    html('<div class="card">')
    html('<p class="sec-hdr">🏢 Fornecedor</p>')
    nf1, nf2 = st.columns(2)
    with nf1:
        forn_opcoes = ["— Fornecedor novo —"] + lista_forn
        forn_sel = st.selectbox("Fornecedor existente", forn_opcoes, index=0, key="np_forn_sel")
        n_forn_novo = st.text_input("Ou novo fornecedor", placeholder="Razão Social completa", key="np_forn_novo")
    with nf2:
        n_nfantasia = st.text_input("Nome Fantasia", placeholder="Opcional", key="np_nf")
        # Auto CNPJ
        _cnpj_hint = ""
        if forn_sel != "— Fornecedor novo —" and not n_forn_novo.strip():
            _cnpj_hint = maps["forn_cnpj"].get(forn_sel, "")
            if _cnpj_hint:
                html(f'<p class="field-auto">🔗 CNPJ auto: {_cnpj_hint}</p>')
        n_cnpj = st.text_input("CNPJ / CPF *", value=_cnpj_hint, placeholder="00.000.000/0001-00", key="np_cnpj")
    n_fornecedor = n_forn_novo.strip() or (forn_sel if forn_sel != "— Fornecedor novo —" else "")
    html('</div>')

    # ── Órgão
    html('<div class="card">')
    html('<p class="sec-hdr">🏛️ Órgão e Secretaria</p>')
    no1, no2 = st.columns(2)
    with no1:
        orgao_opcoes = ["— Órgão novo —"] + lista_orgaos
        orgao_sel = st.selectbox("Órgão existente", orgao_opcoes, index=0, key="np_orgao_sel")
        n_orgao_novo = st.text_input("Ou novo órgão", placeholder="Nome completo do órgão", key="np_orgao_novo")
    with no2:
        _orgao_escolhido = n_orgao_novo.strip() or (orgao_sel if orgao_sel != "— Órgão novo —" else "")
        _sec_auto2 = maps["orgao_sec"].get(_orgao_escolhido, "")
        if _sec_auto2:
            html(f'<p class="field-auto">🔗 Secretário auto: {_sec_auto2}</p>')
        n_secretario = st.text_input("Secretário(a)", value=_sec_auto2, placeholder="Nome do secretário", key="np_sec")
    n_orgao = _orgao_escolhido
    html('</div>')

    # ── Licitação
    html('<div class="card">')
    html('<p class="sec-hdr">📜 Licitação</p>')
    nl1, nl2, nl3 = st.columns(3)
    mod_opcoes = ["— Selecione —"] + lista_mods
    mod_sel    = nl1.selectbox("Modalidade", mod_opcoes, index=0, key="np_mod_sel")
    n_mod_novo = nl1.text_input("Ou nova modalidade", placeholder="ex: DISPENSA 010/2025", key="np_mod_novo")
    _cont_hint = maps["mod_cont"].get(mod_sel, "") if mod_sel != "— Selecione —" else ""
    if _cont_hint: nl2.caption(f"🔗 Sugerido: {_cont_hint}")
    n_contrato     = nl2.text_input("Nº do Contrato",     value=_cont_hint, placeholder="ex: 047/2025", key="np_cont")
    n_ordem_compra = nl3.text_input("N° Ordem de Compra", placeholder="Opcional",                       key="np_oc")
    n_modalidade = n_mod_novo.strip() or (mod_sel if mod_sel != "— Selecione —" else "")
    html('</div>')

    # ── Documento Fiscal
    html('<div class="card">')
    html('<p class="sec-hdr">🧾 Documento Fiscal</p>')
    nd1, nd2, nd3, nd4, nd5 = st.columns([1.5, 1, 1.5, 1.5, 1.5])
    n_tipdoc   = nd1.text_input("Tipo Doc. Fiscal",  placeholder="ex: NFS-e",    key="np_docf")
    n_nf       = nd2.text_input("Nº NF",             placeholder="ex: 229",       key="np_nf_n")
    n_tipnf    = nd3.text_input("Tipo NF",            placeholder="ex: ÚNICO",     key="np_tnf")
    n_valor    = nd4.text_input("Valor (R$) *",       placeholder="ex: 43.088,62", key="np_valor")
    n_periodo  = nd5.text_input("Período de Ref.",    placeholder="ex: OUT/2025",  key="np_per")
    n_data_nf  = st.text_input("Data da NF",          placeholder="ex: 07/10/2025",key="np_dnf")
    html('</div>')

    # ── Objeto
    html('<div class="card">')
    html('<p class="sec-hdr">📝 Objeto da Despesa</p>')
    obj_opcoes = ["— Reutilizar objeto anterior —"] + lista_objs[:60]
    obj_sel    = st.selectbox("Objeto anterior (reutilizar)", obj_opcoes, index=0, key="np_obj_sel")
    n_objeto_novo = st.text_area("Ou descreva *", placeholder="PRESTAÇÃO DE SERVIÇOS DE...", height=72, key="np_obj_novo")
    n_objeto = n_objeto_novo.strip() or (obj_sel if obj_sel != "— Reutilizar objeto anterior —" else "")
    html('</div>')

    # ── Salvar
    html("""<div style="background:#f8fafc;border:1.5px solid #e2e8f0;border-radius:12px;padding:16px 20px;
                        display:flex;align-items:center;gap:12px">
      <div style="flex:1;font-size:13px;color:#64748b">
        <b style="color:#0f172a">Pronto para salvar?</b><br>
        Campos obrigatórios: Nº, Fornecedor, CNPJ, Órgão, Valor, Objeto.
      </div>""")
    _, bsav = st.columns([2, 1])
    with bsav:
        salvar = st.button("💾 Salvar na Planilha", type="primary", use_container_width=True, key="np_salvar")
    html('</div>')
    html('</div>')  # fecha padding

    if salvar:
        _num_v = st.session_state.get("np_num_inp", "").strip()
        erros = []
        if not _num_v:            erros.append("Nº do Documento")
        if not n_fornecedor:      erros.append("Fornecedor")
        if not n_cnpj.strip():    erros.append("CNPJ / CPF")
        if not n_orgao:           erros.append("Órgão / Secretaria")
        if not n_valor.strip():   erros.append("Valor")
        if not n_objeto:          erros.append("Objeto")

        if erros:
            for e in erros: st.error(f"❌ Campo obrigatório: {e}")
        else:
            with st.spinner("Salvando..."):
                try:
                    nova_linha = {
                        "OBJETO":               n_objeto,
                        "ORGÃO":                n_orgao,
                        "MODALIDADE":           n_modalidade,
                        "CONTRATO":             n_contrato,
                        "FORNECEDOR":           n_fornecedor,
                        "NOME FANTASIA":        n_nfantasia,
                        "CNPJ":                 n_cnpj.strip(),
                        "DOCUMENTO FISCAL":     n_tipdoc,
                        "Nº":                   n_nf,
                        "TIPO":                 n_tipnf,
                        "VALOR":                n_valor.strip(),
                        "NÚMERO DO DOCUMENTO":  _num_v,
                        "DATA":                 f"{n_data.day:02d}/{n_data.month:02d}/{n_data.year}",
                        "SECRETARIO":           n_secretario,
                        "N° ORDEM DE COMPRA":   n_ordem_compra,
                        "DATA NF":              n_data_nf,
                        "PERÍODO DE REFERÊNCIA":n_periodo,
                        "DATA PUB.":            f"{n_data_pub.day:02d}/{n_data_pub.month:02d}/{n_data_pub.year}",
                    }
                    if USAR_GSHEETS:
                        salvar_gsheets_linha(nova_linha)
                    else:
                        wb_s = openpyxl.load_workbook(EXCEL_PATH)
                        ws_s = wb_s.active
                        hdrs = [str(ws_s.cell(1, c).value or "").strip() for c in range(1, ws_s.max_column + 1)]
                        if not any(hdrs):
                            ws_s.append(list(nova_linha.keys())); hdrs = list(nova_linha.keys())
                        ws_s.append([nova_linha.get(h, "") for h in hdrs])
                        wb_s.save(EXCEL_PATH)

                    salvar_historico({"processo": _num_v, "orgao": n_orgao,
                                      "fornecedor": n_fornecedor, "valor": n_valor,
                                      "tipo_label": "Cadastro Manual", "decisao": "INCLUÍDO"})
                    st.cache_data.clear()
                    log.info(f"NOVO PROCESSO: #{_num_v} {n_fornecedor}")
                    st.session_state["np_saved"]      = True
                    st.session_state["np_saved_data"] = {"num": _num_v, "forn": n_fornecedor,
                                                          "orgao": n_orgao, "valor": n_valor}
                    st.rerun()
                except PermissionError:
                    st.error("❌ Arquivo em uso! Feche o Excel antes de salvar.")
                except Exception as e:
                    log.error(f"Erro salvar: {e}")
                    st.error(f"❌ {e}")

# ─────────────────────────────────────────────────────────────────────────────
# ══ PÁGINA: BUSCAR / EDITAR ══════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "🔍  Buscar / Editar":
    page_header("🔍", "Buscar / Editar Processo",
                "Pesquise por número, fornecedor, órgão ou objeto", "#0f766e")

    with st.spinner("Carregando..."):
        df = load_df()

    if df.empty:
        st.warning("⚠️ Planilha não encontrada.")
        st.stop()

    html('<div style="padding:0 32px">')

    # Busca imediata: mostra recentes sem digitar
    html('<div class="card">')
    bc1, bc2, bc3 = st.columns([2, 2, 1])
    q_busca  = bc1.text_input("🔎 Buscar (Nº, fornecedor, órgão, objeto...)", placeholder="Digite qualquer termo...", key="busca_q")
    filtro_d = bc2.date_input("Filtrar por data (opcional)", value=None, key="busca_data", format="DD/MM/YYYY")
    with bc3:
        st.write("")
        if st.button("Buscar", type="primary", use_container_width=True, key="btn_buscar"):
            pass

    # Filtro
    dff = df.copy()
    if q_busca.strip():
        mask = pd.Series([False] * len(dff))
        for col in ["NÚMERO DO DOCUMENTO","FORNECEDOR","ORGÃO","OBJETO","CNPJ"]:
            if col in dff.columns:
                mask |= dff[col].astype(str).str.contains(q_busca.strip(), case=False, na=False)
        dff = dff[mask]
        html(f'<div style="font-size:12px;color:#64748b;margin-top:8px"><b>{len(dff)}</b> resultado(s)</div>')
    else:
        # Sem filtro: mostra 15 mais recentes
        dff = dff.tail(15).iloc[::-1]
        html('<div style="font-size:12px;color:#64748b;margin-top:8px">⏱️ 15 processos mais recentes</div>')

    cols_show = [c for c in ["NÚMERO DO DOCUMENTO","ORGÃO","FORNECEDOR","VALOR","DATA","CONTRATO"] if c in dff.columns]
    st.dataframe(dff[cols_show], use_container_width=True, hide_index=True, height=280,
                 column_config={"NÚMERO DO DOCUMENTO": st.column_config.TextColumn("Nº", width="small"),
                                "VALOR": st.column_config.TextColumn("Valor")})
    html('</div>')

    # Selecionar processo para ver detalhes
    if not dff.empty:
        html('<p class="sec-hdr">Selecione para ver / editar</p>')
        nums_disp = ["— selecione —"] + dff["NÚMERO DO DOCUMENTO"].unique().tolist()
        num_ed = st.selectbox("", nums_disp, label_visibility="collapsed", key="busca_sel")

        if num_ed != "— selecione —":
            row = df[df["NÚMERO DO DOCUMENTO"].astype(str) == str(num_ed)].iloc[0]
            html(f"""
            <div style="background:#f0fdf4;border:1.5px solid #86efac;border-radius:10px;
                        padding:12px 16px;margin:12px 0">
              <span style="font-weight:700;color:#15803d">✓ Processo #{num_ed} carregado</span>
              <span style="color:#166534;font-size:13px;margin-left:12px">
                {row.get('FORNECEDOR','')} &nbsp;·&nbsp; {row.get('ORGÃO','')} &nbsp;·&nbsp; {row.get('VALOR','')}
              </span>
            </div>""")

            with st.expander("📋 Ver todos os campos", expanded=True):
                campos_linha = row.to_dict()
                col_a, col_b = st.columns(2)
                keys = [k for k in campos_linha.keys()]
                for i, k in enumerate(keys):
                    c = col_a if i % 2 == 0 else col_b
                    c.text_input(k, value=str(campos_linha[k]), key=f"ed_{k}_{num_ed}", disabled=True)

            ba, bb, _ = st.columns([1, 1, 2])
            with ba:
                if st.button("⚡ Gerar Documentos", type="primary", use_container_width=True):
                    st.session_state["proc_sel"] = row.to_dict()
                    st.session_state["tipo"]     = sugerir_tipo(row.get("OBJETO",""))
                    st.session_state["etapa"]    = 2
                    # Navega para Gerar Documentos
                    st.info("👆 Vá para ⚡ Gerar Documentos no menu lateral para continuar.")
    html('</div>')

# ─────────────────────────────────────────────────────────────────────────────
# ══ PÁGINA: HISTÓRICO ════════════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "📊  Histórico":
    page_header("📊", "Histórico de Documentos", "Todos os documentos gerados pelo sistema", "#7c3aed")
    html('<div style="padding:0 32px 32px">')

    # Aba Logs do Sistema
    tab_hist, tab_logs = st.tabs(["📋 Histórico de Geração", "🔍 Logs do Sistema"])

    with tab_hist:
        if USAR_GSHEETS:
            st.info("☁️ Histórico disponível apenas no modo Excel local.")
        elif not os.path.exists(EXCEL_PATH):
            st.warning("⚠️ Planilha não encontrada.")
        else:
            try:
                wb = openpyxl.load_workbook(EXCEL_PATH)
                if "Histórico" not in wb.sheetnames:
                    st.info("💡 Nenhum documento gerado ainda.")
                else:
                    ws = wb["Histórico"]
                    data = list(ws.values)
                    if len(data) > 1:
                        hist_df = pd.DataFrame(data[1:], columns=data[0])
                        html(f'<p style="font-size:12px;color:#64748b;margin-bottom:8px"><b>{len(hist_df)}</b> documento(s) gerado(s)</p>')
                        # Filtro rápido
                        q_h = st.text_input("🔎 Filtrar histórico", placeholder="Processo, operador, tipo...", key="hist_q")
                        if q_h:
                            mask_h = hist_df.astype(str).apply(lambda col: col.str.contains(q_h, case=False, na=False)).any(axis=1)
                            hist_df = hist_df[mask_h]
                        st.dataframe(hist_df.iloc[::-1], use_container_width=True, hide_index=True, height=400)
            except Exception as e:
                st.error(f"Erro: {e}")

    with tab_logs:
        html('<p class="sec-hdr">📋 Log mais recente (sistema.log)</p>')
        log_path = os.path.join(_BASE, "sistema.log")
        if os.path.exists(log_path):
            try:
                linhas = open(log_path, encoding="utf-8").readlines()[-100:]
                log_text = "".join(reversed(linhas))
                st.code(log_text, language="text")
            except Exception as e:
                st.error(f"Erro ao ler log: {e}")
        else:
            st.info("Nenhum log registrado ainda.")

        if st.button("🗑️ Limpar Log", key="limpar_log"):
            try:
                open(log_path, "w").close()
                st.success("Log limpo.")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    html('</div>')

# ─────────────────────────────────────────────────────────────────────────────
# ══ PÁGINA: USUÁRIOS ═════════════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "👥  Usuários":
    if not _is_admin:
        st.error("🔒 Acesso restrito a administradores.")
        st.stop()

    page_header("👥", "Gestão de Usuários", "Criar, editar, desativar e redefinir senhas", "#dc2626")
    html('<div style="padding:0 32px 32px">')

    tab_u1, tab_u2, tab_u3 = st.tabs(["👥 Usuários", "➕ Criar", "🔑 Minha Senha"])

    with tab_u1:
        users = _carregar_usuarios()
        for login, u in users.items():
            ativo  = u.get("ativo", True)
            perfil = u.get("perfil", "operador")
            badge_cor = {"admin": "badge-red", "gestor": "badge-purple", "operador": "badge-blue"}.get(perfil, "badge-blue")
            html(f"""
            <div class="card" style="padding:14px 18px;margin-bottom:10px">
              <div style="display:flex;align-items:center;gap:12px">
                <div style="font-size:28px">{"👑" if perfil=="admin" else "🎯" if perfil=="gestor" else "👤"}</div>
                <div style="flex:1">
                  <div style="font-weight:700;font-size:14px;color:#0f172a">{u.get('nome',login)}</div>
                  <div style="font-size:12px;color:#64748b">@{login}</div>
                </div>
                <span class="badge {badge_cor}">{perfil}</span>
                <span class="badge {'badge-green' if ativo else 'badge-red'}">{'Ativo' if ativo else 'Inativo'}</span>
              </div>
            </div>""")
            ca, cb, cc = st.columns([1, 1, 1])
            with ca:
                lbl_ativo = "🔴 Desativar" if ativo else "🟢 Ativar"
                if st.button(lbl_ativo, key=f"toggle_{login}", use_container_width=True):
                    users = _carregar_usuarios()
                    users[login]["ativo"] = not ativo
                    _salvar_usuarios(users)
                    st.rerun()
            with cb:
                if login != st.session_state.get("user") and st.button("🗑️ Excluir", key=f"del_{login}", use_container_width=True):
                    if f"confirm_del_{login}" not in st.session_state:
                        st.session_state[f"confirm_del_{login}"] = True
                    else:
                        users = _carregar_usuarios()
                        del users[login]
                        _salvar_usuarios(users)
                        st.success(f"Usuário {login} excluído.")
                        st.rerun()
                if st.session_state.get(f"confirm_del_{login}"):
                    cc.warning(f"Clique Excluir novamente para confirmar.")

    with tab_u2:
        html('<div class="card">')
        with st.form("form_criar"):
            nu1, nu2 = st.columns(2)
            n_login  = nu1.text_input("Login *", placeholder="ex: joao.silva")
            n_nome   = nu2.text_input("Nome Completo *")
            n_senha  = nu1.text_input("Senha *", type="password", placeholder="Mínimo 6 chars")
            n_conf   = nu2.text_input("Confirmar Senha *", type="password")
            n_perf   = st.selectbox("Perfil", ["operador", "gestor", "admin"])
            if st.form_submit_button("✅ Criar Usuário", type="primary"):
                if not n_login.strip() or not n_nome.strip() or not n_senha.strip():
                    st.error("Preencha todos os campos.")
                elif n_senha != n_conf:
                    st.error("Senhas não coincidem.")
                elif len(n_senha) < 6:
                    st.error("Senha deve ter mínimo 6 caracteres.")
                elif n_login.strip() in _carregar_usuarios():
                    st.error("Login já existe.")
                else:
                    salt = secrets.token_hex(16)
                    h    = hashlib.sha256((salt + n_senha).encode()).hexdigest()
                    users = _carregar_usuarios()
                    users[n_login.strip()] = {"senha": h, "salt": salt, "nome": n_nome.strip(),
                                               "perfil": n_perf, "ativo": True}
                    _salvar_usuarios(users)
                    st.success(f"✅ Usuário @{n_login} criado!")
        html('</div>')

    with tab_u3:
        html('<div class="card">')
        with st.form("form_senha"):
            s_atual = st.text_input("Senha Atual *", type="password")
            s_nova  = st.text_input("Nova Senha *",  type="password")
            s_conf  = st.text_input("Confirmar Nova Senha *", type="password")
            if st.form_submit_button("💾 Alterar Senha", type="primary"):
                u_ck = _verificar_senha(st.session_state["user"], s_atual)
                if not u_ck:
                    st.error("Senha atual incorreta.")
                elif s_nova != s_conf:
                    st.error("Novas senhas não coincidem.")
                elif len(s_nova) < 6:
                    st.error("Mínimo 6 caracteres.")
                else:
                    salt = secrets.token_hex(16)
                    h    = hashlib.sha256((salt + s_nova).encode()).hexdigest()
                    users = _carregar_usuarios()
                    users[st.session_state["user"]].update({"senha": h, "salt": salt})
                    _salvar_usuarios(users)
                    st.success("✅ Senha alterada com sucesso!")
        html('</div>')

    html('</div>')

# ─────────────────────────────────────────────────────────────────────────────
# ══ PÁGINA: CONFIGURAÇÕES ════════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────
elif pagina == "⚙️  Configurações":
    page_header("⚙️", "Configurações", "Planilha, sincronização e sistema")
    html('<div style="padding:0 32px 32px">')

    tab_c1, tab_c2, tab_c3 = st.tabs(["📂 Planilha", "☁️ Google Sheets", "ℹ️ Sistema"])

    with tab_c1:
        html('<div class="card">')
        html('<p class="sec-hdr">Caminho da planilha Excel</p>')
        novo_path = st.text_input("Caminho", value=EXCEL_PATH, key="cfg_path")
        cg1, cg2, _ = st.columns([1, 1, 2])
        with cg1:
            if st.button("💾 Salvar", type="primary", use_container_width=True):
                try:
                    cfg_p = os.path.join(_BASE, "config.txt")
                    lines = open(cfg_p, encoding="utf-8").readlines() if os.path.exists(cfg_p) else []
                    novas = [l for l in lines if not l.strip().startswith("PLANILHA")]
                    novas.append(f"PLANILHA = {novo_path.strip()}\n")
                    open(cfg_p, "w", encoding="utf-8").writelines(novas)
                    st.success("✅ Salvo! Reinicie o sistema."); st.cache_data.clear()
                except Exception as e:
                    st.error(str(e))
        with cg2:
            if st.button("🔄 Testar Conexão", use_container_width=True):
                if os.path.exists(novo_path.strip()):
                    df_t = pd.read_excel(novo_path.strip(), dtype=str)
                    st.success(f"✅ Planilha OK — {len(df_t)} linha(s)")
                else:
                    st.error(f"❌ Arquivo não encontrado: {novo_path}")
        html('</div>')

    with tab_c2:
        html('<div class="card">')
        st.info("""
        **☁️ Google Sheets — Sincronização em tempo real**

        Para ativar o modo nuvem, configure as credenciais no arquivo `.streamlit/secrets.toml`:
        ```toml
        [gcp_service_account]
        type = "service_account"
        project_id = "seu-projeto"
        private_key_id = "..."
        private_key = "-----BEGIN PRIVATE KEY-----..."
        client_email = "..."

        [google_sheets]
        sheet_id = "id-da-sua-planilha"
        ```
        E no `config.txt`, altere para: `USAR_GSHEETS = sim`
        """)
        modo_atual = "☁️ Google Sheets" if USAR_GSHEETS else "💾 Excel Local"
        st.metric("Modo Atual", modo_atual)
        if USAR_GSHEETS:
            if st.button("🔄 Testar Google Sheets", type="primary"):
                with st.spinner("Testando conexão..."):
                    try:
                        df_t = load_df()
                        st.success(f"✅ Conectado! {len(df_t)} linha(s) carregadas.")
                    except Exception as e:
                        st.error(f"❌ {e}")
        html('</div>')

    with tab_c3:
        html('<div class="card">')
        infos = {
            "Versão":          "5.0 — Light Mode Moderno",
            "Framework":       "Streamlit (Web-based, multi-dispositivo)",
            "Motor PDF":       "ReportLab (cross-platform Windows/Linux)",
            "Autenticação":    "SHA-256 + Salt (arquivo local usuarios.json)",
            "Banco de Dados":  f"Excel: {os.path.basename(EXCEL_PATH)}" if not USAR_GSHEETS else "Google Sheets (nuvem)",
            "Logs":            "sistema.log (rastreamento completo de erros)",
            "Multiusuário":    "Admin / Gestor / Operador",
            "Servidor":        f"{get_ip()}:8501",
        }
        for k, v in infos.items():
            cl, cv = st.columns([1, 2])
            cl.markdown(f'<span style="font-size:12px;font-weight:600;color:#374151">{k}</span>', unsafe_allow_html=True)
            cv.markdown(f'<span style="font-size:12px;color:#64748b">{v}</span>', unsafe_allow_html=True)

        html('<div style="margin-top:16px"></div>')
        if st.button("🗑️ Limpar Cache", use_container_width=False):
            st.cache_data.clear(); st.success("Cache limpo!"); st.rerun()
        html('</div>')

    html('</div>')
